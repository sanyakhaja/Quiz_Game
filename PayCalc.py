import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook

class PayrollApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Payroll Calculator")

        self.total_hours = 0
        self.total_pay = 0
        self.tax_amount = 0

        self.load_excel_button = ttk.Button(root, text="Load Excel", command=self.load_excel)
        self.load_excel_button.pack(pady=10)

        self.result_label = tk.Label(root, text="Results:")
        self.result_label.pack()

        self.hours_label = tk.Label(root, text="")
        self.hours_label.pack()

        self.pay_label = tk.Label(root, text="")
        self.pay_label.pack()

        self.tax_label = tk.Label(root, text="")
        self.tax_label.pack()

    def load_excel(self):
        try:
            workbook = load_workbook("excel-sheet.xlsx")
            sheet = workbook.active

            self.total_hours = 0
            self.total_pay = 0

            for row in sheet.iter_rows(values_only=True):
                if len(row) >= 2:
                    hours, pay = row[0], row[1]
                    self.total_hours += hours
                    self.total_pay += pay

            self.tax_amount = self.total_pay * 0.2

            self.update_results()
        except Exception as e:
            self.result_label.config(text="Error loading Excel sheet")

    def update_results(self):
        self.result_label.config(text="Results:")
        self.hours_label.config(text=f"Total Hours Worked: {self.total_hours}")
        self.pay_label.config(text=f"Total Pay Accumulated: {self.total_pay}")
        self.tax_label.config(text=f"Tax Amount to Save (20%): {self.tax_amount:.2f}")

if __name__ == "__main__":
    root = tk.Tk()
    app = PayrollApp(root)
    root.mainloop()
